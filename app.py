import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Sayfa Yapılandırması
st.set_page_config(
    page_title="Tekstilbox & EasyExpo B2B Lead Finder",
    page_icon="🎯",
    layout="wide"
)

# KULLANICI GİRİŞ KONTROLÜ
AUTHORIZED_USERS = {
    "admin": "tekstilbox2026",
    "satis1": "satis2026",
    "satis2": "ees2026"
}

def check_password():
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    if st.session_state["password_correct"]:
        return True

    st.subheader("🔑 Tekstilbox & EasyExpo - B2B Lead Finder Giriş Paneli")
    col1, col2 = st.columns(2)
    with col1:
        username = st.text_input("Kullanıcı Adı")
        password = st.text_input("Şifre", type="password")
        if st.button("Giriş Yap 🚀"):
            if username in AUTHORIZED_USERS and AUTHORIZED_USERS[username] == password:
                st.session_state["password_correct"] = True
                st.rerun()
            else:
                st.error("❌ Hatalı kullanıcı adı veya şifre!")
    return False

# Web sitesi bulma fonksiyonu
def find_company_website(company_name, country_code, api_key):
    try:
        search_url = "https://serpapi.com/search"
        params = {
            "engine": "google",
            "q": f"{company_name} official website {country_code}",
            "gl": country_code.lower(),
            "num": "3",
            "api_key": api_key
        }
        res = requests.get(search_url, params=params).json()
        results = res.get("organic_results", [])
        for r in results:
            link = r.get("link", "")
            if not any(x in link for x in ["linkedin.com", "facebook.com", "instagram.com", "twitter.com", "youtube.com", "yelp.", "tripadvisor."]):
                return link.split("/")[2] if "://" in link else link
    except:
        pass
    return "LinkedIn sayfasından manuel kontrol edin"

if check_password():
    st.sidebar.image("https://www.tekstilbox.com/wp-content/uploads/2021/04/cropped-tekstilbox-logo-black-1.png", width=200)
    st.sidebar.title("⚙️ Ayarlar")
    
    serpapi_key = st.sidebar.text_input(
        "SerpApi Anahtarınız (API Key)", 
        type="password"
    )
    
    st.sidebar.markdown("---")
    st.sidebar.info("Tüm Avrupa Sürümü: 40+ Avrupa ülkesi, manuel kelime girişi, 100 sayfa tarama ve gelişmiş dil kütüphanesi aktif.")

    st.title("🎯 Tekstilbox & EasyExpo B2B Arama Motoru")

    # Tüm Avrupa Ülkeleri ve LinkedIn Subdomain Eşleştirmeleri
    EUROPEAN_COUNTRIES = {
        "Almanya (DE)": {"code": "DE", "domain": "de.linkedin.com/company/"},
        "Avusturya (AT)": {"code": "AT", "domain": "at.linkedin.com/company/"},
        "Arnavutluk (AL)": {"code": "AL", "domain": "al.linkedin.com/company/"},
        "Andorra (AD)": {"code": "AD", "domain": "linkedin.com/company/"},
        "Belçika (BE)": {"code": "BE", "domain": "be.linkedin.com/company/"},
        "Birleşik Krallık / İngiltere (GB)": {"code": "GB", "domain": "uk.linkedin.com/company/"},
        "Bosna Hersek (BA)": {"code": "BA", "domain": "ba.linkedin.com/company/"},
        "Bulgaristan (BG)": {"code": "BG", "domain": "bg.linkedin.com/company/"},
        "Çek Cumhuriyeti (CZ)": {"code": "CZ", "domain": "cz.linkedin.com/company/"},
        "Danimarka (DK)": {"code": "DK", "domain": "dk.linkedin.com/company/"},
        "Estonya (EE)": {"code": "EE", "domain": "ee.linkedin.com/company/"},
        "Finlandiya (FI)": {"code": "FI", "domain": "fi.linkedin.com/company/"},
        "Fransa (FR)": {"code": "FR", "domain": "fr.linkedin.com/company/"},
        "Hırvatistan (HR)": {"code": "HR", "domain": "hr.linkedin.com/company/"},
        "Hollanda (NL)": {"code": "NL", "domain": "nl.linkedin.com/company/"},
        "İrlanda (IE)": {"code": "IE", "domain": "ie.linkedin.com/company/"},
        "İspanya (ES)": {"code": "ES", "domain": "es.linkedin.com/company/"},
        "İsveç (SE)": {"code": "SE", "domain": "se.linkedin.com/company/"},
        "İsviçre (CH)": {"code": "CH", "domain": "ch.linkedin.com/company/"},
        "İtalya (IT)": {"code": "IT", "domain": "it.linkedin.com/company/"},
        "İzlanda (IS)": {"code": "IS", "domain": "is.linkedin.com/company/"},
        "Kosova (XK)": {"code": "XK", "domain": "linkedin.com/company/"},
        "Letonya (LV)": {"code": "LV", "domain": "lv.linkedin.com/company/"},
        "Lihtenştayn (LI)": {"code": "LI", "domain": "linkedin.com/company/"},
        "Litvanya (LT)": {"code": "LT", "domain": "lt.linkedin.com/company/"},
        "Lüksemburg (LU)": {"code": "LU", "domain": "lu.linkedin.com/company/"},
        "Macaristan (HU)": {"code": "HU", "domain": "hu.linkedin.com/company/"},
        "Kuzey Makedonya (MK)": {"code": "MK", "domain": "mk.linkedin.com/company/"},
        "Malta (MT)": {"code": "MT", "domain": "mt.linkedin.com/company/"},
        "Moldova (MD)": {"code": "MD", "domain": "md.linkedin.com/company/"},
        "Monako (MC)": {"code": "MC", "domain": "linkedin.com/company/"},
        "Karadağ (ME)": {"code": "ME", "domain": "me.linkedin.com/company/"},
        "Norveç (NO)": {"code": "NO", "domain": "no.linkedin.com/company/"},
        "Polonya (PL)": {"code": "PL", "domain": "pl.linkedin.com/company/"},
        "Portekiz (PT)": {"code": "PT", "domain": "pt.linkedin.com/company/"},
        "Romanya (RO)": {"code": "RO", "domain": "ro.linkedin.com/company/"},
        "Sırbistan (RS)": {"code": "RS", "domain": "rs.linkedin.com/company/"},
        "Slovakya (SK)": {"code": "SK", "domain": "sk.linkedin.com/company/"},
        "Slovenya (SI)": {"code": "SI", "domain": "si.linkedin.com/company/"},
        "Türkiye (TR)": {"code": "TR", "domain": "tr.linkedin.com/company/"},
        "Ukrayna (UA)": {"code": "UA", "domain": "ua.linkedin.com/company/"},
        "Yunanistan (GR)": {"code": "GR", "domain": "gr.linkedin.com/company/"}
    }

    # TÜM AVRUPA İÇİN GELİŞMİŞ ÇOKLU DİL SÖZLÜĞÜ (40+ Ülke)
    LOCAL_DICTIONARY = {
        "FR": { "shopfitter": "agenceur", "lightbox": "caisson lumineux", "retail design": "architecture commerciale", "shopfitting": "agencement de magasin", "signage": "signalétique", "sign": "enseigne", "profile": "profilé", "fabric": "tissu", "stand": "stand d'exposition" },
        "DE": { "shopfitter": "ladenbauer", "lightbox": "leuchtkasten", "retail design": "ladenbau design", "shopfitting": "ladeneinrichtung", "signage": "beschilderung", "sign": "schilder", "profile": "profile", "fabric": "textilfalwand", "stand": "messestand" },
        "AT": { "shopfitter": "ladenbauer", "lightbox": "leuchtkasten", "retail design": "ladenbau design", "shopfitting": "ladeneinrichtung", "signage": "beschilderung", "sign": "schilder", "profile": "profile", "fabric": "textilfalwand", "stand": "messestand" },
        "CH": { "shopfitter": "ladenbauer", "lightbox": "leuchtkasten", "retail design": "ladenbau design", "shopfitting": "ladeneinrichtung", "signage": "beschilderung", "sign": "schilder", "profile": "profile", "fabric": "textilfalwand", "stand": "messestand" },
        "NL": { "shopfitter": "winkelinterieurbouw", "lightbox": "lichtbak", "retail design": "retail design", "shopfitting": "winkelindeling", "signage": "bewegwijzering", "sign": "borden", "profile": "profiel", "fabric": "stof", "stand": "beursstand" },
        "BE": { "shopfitter": "winkelinterieurbouw", "lightbox": "lichtbak", "retail design": "retail design", "shopfitting": "winkelindeling", "signage": "bewegwijzering", "sign": "borden", "profile": "profiel", "fabric": "stof", "stand": "beursstand" },
        "IT": { "shopfitter": "allestitore", "lightbox": "cassonetto luminoso", "retail design": "progettazione negozi", "shopfitting": "allestimento negozi", "signage": "segnaletica", "sign": "insegna", "profile": "profilo", "fabric": "tessuto", "stand": "stand fieristico" },
        "ES": { "shopfitter": "instalador de tiendas", "lightbox": "caja de luz", "retail design": "diseño comercial", "shopfitting": "equipamiento comercial", "signage": "señalización", "sign": "rótulo", "profile": "perfil", "fabric": "tela", "stand": "stand ferial" },
        "PT": { "shopfitter": "instalador de lojas", "lightbox": "caixa de luz", "retail design": "design de retalho", "shopfitting": "equipamento comercial", "signage": "sinalização", "sign": "letreiro", "profile": "perfil", "fabric": "tecido", "stand": "stand de feira" },
        "PL": { "shopfitter": "wyposażenie sklepów", "lightbox": "kaseton świetlny", "retail design": "projektowanie sklepów", "shopfitting": "aranżacja sklepów", "signage": "oznakowanie", "sign": "szyld", "profile": "profil", "fabric": "tkanina", "stand": "stoisko targowe" },
        "TR": { "shopfitter": "mağaza dekorasyon", "lightbox": "ışıklı tabela", "retail design": "mağaza tasarımı", "shopfitting": "mağaza donanımları", "signage": "yönlendirme tabelaları", "sign": "tabela", "profile": "profil", "fabric": "kumaş", "stand": "fuar standı" },
        "LV": { "shopfitter": "veikalu aprīkojums", "lightbox": "gaismas kastes", "retail design": "mazumtirdzniecības dizains", "shopfitting": "veikalu iekārtošana", "signage": "izkārtnes", "sign": "izkārtne", "profile": "profils", "fabric": "audums", "stand": "stends" },
        "SE": { "shopfitter": "butiksinredning", "lightbox": "ljuslåda", "retail design": "butiksdesign", "shopfitting": "butiksutrustning", "signage": "skyltning", "sign": "skylt", "profile": "profil", "fabric": "tyg", "stand": "monter" },
        "NO": { "shopfitter": "butikkinnredning", "lightbox": "lyskasse", "retail design": "butikkdesign", "shopfitting": "butikkutstyr", "signage": "skilting", "sign": "skilt", "profile": "profil", "fabric": "stoff", "stand": "messestand" },
        "DK": { "shopfitter": "butiksinventar", "lightbox": "lyskasse", "retail design": "butiksdesign", "shopfitting": "butiksindretning", "signage": "skiltning", "sign": "skilt", "profile": "profil", "fabric": "stof", "stand": "messebod" },
        "FI": { "shopfitter": "myymäläkalusteet", "lightbox": "valokaappi", "retail design": "myymäläsuunnittelu", "shopfitting": "myymäläkalustus", "signage": "opasteet", "sign": "kyltti", "profile": "profiili", "fabric": "kangas", "stand": "messuosasto" },
        "RO": { "shopfitter": "amenajari comerciale", "lightbox": "caseta luminoasa", "retail design": "design retail", "shopfitting": "mobilier comercial", "signage": "semnalistica", "sign": "firma luminoasa", "profile": "profil", "fabric": "material textil", "stand": "stand expozitional" },
        "CZ": { "shopfitter": "vybavení obchodů", "lightbox": "světelný box", "retail design": "retail design", "shopfitting": "vybavení prodejen", "signage": "značení", "sign": "nápis", "profile": "profil", "fabric": "látka", "stand": "výstavní stánek" },
        "SK": { "shopfitter": "vybavenie obchodov", "lightbox": "svetelný box", "retail design": "dizajn predajní", "shopfitting": "zariadenie predajní", "signage": "značenie", "sign": "nápis", "profile": "profil", "fabric": "látka", "stand": "výstavný stánok" },
        "HU": { "shopfitter": "üzletberendezés", "lightbox": "világítótábla", "retail design": "üzlettervezés", "shopfitting": "üzletkialakítás", "signage": "cégér", "sign": "tábla", "profile": "profil", "fabric": "szövet", "stand": "kiállítási stand" },
        "BG": { "shopfitter": "обзавеждане за магазини", "lightbox": "светеща кутия", "retail design": "ритейл дизайн", "shopfitting": "интериор на магазин", "signage": "сигналистика", "sign": "табела", "profile": "профил", "fabric": "плат", "stand": "изложбен щанд" },
        "GR": { "shopfitter": "εξοπλισμός καταστημάτων", "lightbox": "φωτεινή επιγραφή", "retail design": "σχεδιασμός καταστημάτων", "shopfitting": "επίπλωση καταστημάτων", "signage": "σήμανση", "sign": "πινακίδα", "profile": "προφίλ", "fabric": "ύφασμα", "stand": "εκθεσιακό περίπτερο" },
        "EE": { "shopfitter": "kaupluse sisustus", "lightbox": "valguskast", "retail design": "jaekaubanduse disain", "shopfitting": "kaupluse sisustamine", "signage": "sildid", "sign": "silt", "profile": "profiil", "fabric": "kangas", "stand": "messiboks" },
        "LT": { "shopfitter": "parduotuvių įranga", "lightbox": "šviesdėžė", "retail design": "mažmeninės prekybos dizainas", "shopfitting": "parduotuvių įrengimas", "signage": "iškabos", "sign": "iškaba", "profile": "profilis", "fabric": "audinys", "stand": "parodų stendas" },
        "HR": { "shopfitter": "opremanje trgovina", "lightbox": "svjetleća reklama", "retail design": "dizajn interijera trgovina", "shopfitting": "uređenje trgovine", "signage": "oznake", "sign": "natpis", "profile": "profil", "fabric": "tkanina", "stand": "izložbeni štand" },
        "BA": { "shopfitter": "opremanje trgovina", "lightbox": "svjetleća reklama", "retail design": "dizajn interijera trgovina", "shopfitting": "uređenje trgovine", "signage": "oznake", "sign": "natpis", "profile": "profil", "fabric": "tkanina", "stand": "izložbeni štand" },
        "RS": { "shopfitter": "opremanje prodavnica", "lightbox": "svetleća reklama", "retail design": "dizajn enterijera", "shopfitting": "uređenje lokala", "signage": "oznake", "sign": "natpis", "profile": "profil", "fabric": "tkanina", "stand": "izložbeni štand" },
        "SI": { "shopfitter": "oprema trgovin", "lightbox": "svetlobna tabla", "retail design": "oblikovanje trgovin", "shopfitting": "ureditev trgovine", "signage": "označevanje", "sign": "napis", "profile": "profil", "fabric": "tkanina", "stand": "razstavni prostor" },
        "UA": { "shopfitter": "торгове обладнання", "lightbox": "лайтбокс", "retail design": "дизайн магазину", "shopfitting": "облаштування магазину", "signage": "вивіски", "sign": "вивіска", "profile": "профіль", "fabric": "тканина", "stand": "виставковий стенд" },
        "IS": { "shopfitter": "verslanainnréttingar", "lightbox": "ljósakassi", "retail design": "verslanahönnun", "shopfitting": "verslanabúnaður", "signage": "skilti", "sign": "skilti", "profile": "prófíl", "fabric": "efni", "stand": "sýningarbás" },
        "MK": { "shopfitter": "опрема за продавници", "lightbox": "светлечка реклама", "retail design": "дизајн на продавници", "shopfitting": "уредување на продавници", "signage": "ознаки", "sign": "натпис", "profile": "профил", "fabric": "ткаенина", "stand": "изложбен штанд" }
    }

    col_left, col_right = st.columns([1, 2])

    with col_left:
        # 1. Ülke Seçimi
        selected_country_name = st.selectbox(
            "1. Hedef Ülke Seçin", 
            list(EUROPEAN_COUNTRIES.keys())
        )
        country_info = EUROPEAN_COUNTRIES[selected_country_name]
        country_code = country_info["code"]
        linkedin_subdomain = country_info["domain"]

        # 2. Hedef Kitle Seçimi
        hedef_kitle = st.selectbox(
            "2. Hedef Kitle Seçin", 
            ["Shopfitter", "Signage/Sign", "Producer", "Print house", "Modular Exhibition Systems"]
        )

        keywords_to_search = []

        if hedef_kitle == "Shopfitter":
            keywords_to_search = st.multiselect("Aranacak Anahtar Kelimeler", ["shopfitter", "lightbox", "retail design", "shopfitting"], default=["shopfitter"])
        elif hedef_kitle == "Signage/Sign":
            keywords_to_search = st.multiselect("Aranacak Anahtar Kelimeler", ["lightbox", "signage", "sign"], default=["lightbox"])
        elif hedef_kitle == "Producer":
            producer_type = st.radio("Ürün Grubu Seçin", ["Profil", "Kumaş", "LED"])
            if producer_type == "Profil":
                keywords_to_search = st.multiselect("Profil Seçenekleri", ["double side", "single side", "lightbox", "profile"])
            elif producer_type == "Kumaş":
                keywords_to_search = st.multiselect("Kumaş Seçenekleri", ["SEG Fabric", "UV Fabric", "Dye-Sub Fabric", "Sublimation Fabric"])
            elif producer_type == "LED":
                keywords_to_search = st.multiselect("LED Seçenekleri", ["module led", "lightbox led"])
        elif hedef_kitle == "Print house":
            keywords_to_search = st.multiselect("Baskı Teknolojileri Seçin", ["Digital Printing", "UV printing", "Dye-Sub printing", "Sublimation printing", "large format printing"])
        elif hedef_kitle == "Modular Exhibition Systems":
            moduler_type = st.radio("Sistem Seçin", ["xPowall+", "EES"])
            if moduler_type == "xPowall+":
                keywords_to_search = st.multiselect("Marka/Uyum Seçenekleri", ["beMatrix", "Aluvision"])
            elif moduler_type == "EES":
                keywords_to_search = st.multiselect("Sistem Seçenekleri", ["Modular Exhibition Systems", "Modular Expo Systems", "Stand"])

        # 3. YENİ ÖZELLİK: Manuel Kelime Girişi
        st.markdown("---")
        custom_keyword = st.text_input("➕ Kendi Kelimenizi Yazın (Opsiyonel)", placeholder="Örn: display, akrilik, ahşap...")

    with col_right:
        st.subheader("🔍 Arama Bilgileri ve Önizleme")
        
        # Seçilen hazır kelimeler ve kullanıcının yazdığı özel kelimeyi birleştir
        final_keywords = list(keywords_to_search)
        if custom_keyword.strip():
            final_keywords.append(custom_keyword.strip())
        
        if not final_keywords:
            st.warning("Lütfen sol taraftan en az bir hazır anahtar kelime seçin veya kendi kelimenizi yazın.")
        else:
            query_parts = []
            local_dict = LOCAL_DICTIONARY.get(country_code, {})
            
            for kw in final_keywords:
                local_kw = local_dict.get(kw.lower())
                if local_kw and local_kw != kw.lower():
                    query_parts.append(f'("{kw}" OR "{local_kw}")')
                else:
                    query_parts.append(f'"{kw}"')
                    
            search_query = f"site:{linkedin_subdomain} " + " AND ".join(query_parts)
            
            st.code(f"Oluşturulan Google Arama Sorgusu:\n{search_query}", language="text")
            
            if st.button("Müşterileri Bul 🚀", use_container_width=True):
                if not serpapi_key:
                    st.error("⚠️ Lütfen sol menüdeki 'Ayarlar' kısmından SerpApi anahtarınızı girin!")
                else:
                    st.info("🔄 Google verileri derinlemesine taranıyor. Çoklu sayfalar taranacağı için bu işlem 15-30 saniye sürebilir...")
                    
                    leads = []
                    for page in range(0, 3):
                        url = "https://serpapi.com/search"
                        params = {
                            "engine": "google",
                            "q": search_query,
                            "gl": country_code.lower(),
                            "start": str(page * 10),
                            "num": "10",
                            "api_key": serpapi_key
                        }
                        
                        try:
                            response = requests.get(url, params=params)
                            data = response.json()
                            results = data.get("organic_results", [])
                            
                            if not results:
                                break
                                
                            for result in results:
                                linkedin_url = result.get("link", "")
                                if "linkedin.com/company/" in linkedin_url:
                                    raw_title = result.get("title", "Bilinmeyen Firma")
                                    clean_title = raw_title.split(":")[0].split("|")[0].split("-")[0].strip()
                                    
                                    if not any(lead['LinkedIn Sayfası Linki'] == linkedin_url for lead in leads):
                                        leads.append({
                                            "Ülke": country_code,
                                            "Firma İsmi": clean_title,
                                            "Web Sitesi": "Aranıyor...",
                                            "LinkedIn Sayfası Linki": linkedin_url
                                        })
                            time.sleep(1)
                        except Exception as e:
                            st.error(f"Sayfa {page+1} taranırken hata oluştu: {e}")
                            break
                    
                    if leads:
                        st.info("🌐 Bulunan firmaların web siteleri tespit ediliyor, lütfen bekleyin...")
                        progress_bar = st.progress(0)
                        
                        for idx, lead in enumerate(leads):
                            website = find_company_website(lead["Firma İsmi"], country_code, serpapi_key)
                            lead["Web Sitesi"] = website
                            progress_bar.progress((idx + 1) / len(leads))
                            time.sleep(0.5)
                        
                        df = pd.DataFrame(leads)
                        st.success(f"🎉 Başarılı! Toplam {len(df)} adet potansiyel firma listelendi.")
                        st.dataframe(df, use_container_width=True)
                        
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name="Müşteri Listesi")
                            
                            workbook = writer.book
                            worksheet = writer.sheets["Müşteri Listesi"]
                            
                            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                            center_align = Alignment(horizontal="center", vertical="center")
                            left_align = Alignment(horizontal="left", vertical="center")
                            
                            thin_border = Border(
                                left=Side(style='thin', color='D3D3D3'),
                                right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'),
                                bottom=Side(style='thin', color='D3D3D3')
                            )
                            
                            for col_idx, col in enumerate(df.columns, 1):
                                cell = worksheet.cell(row=1, column=col_idx)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = center_align
                                cell.border = thin_border
                            
                            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=4):
                                for cell in row:
                                    cell.font = Font(name="Arial", size=10)
                                    cell.border = thin_border
                                    if cell.column == 1:
                                        cell.alignment = center_align
                                    else:
                                        cell.alignment = left_align
                            
                            for col in worksheet.columns:
                                max_len = max(len(str(cell.value or '')) for cell in col)
                                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                                
                        excel_data = output.getvalue()
                        
                        st.download_button(
                            label="📥 Excel Dosyasını İndir (.xlsx)",
                            data=excel_data,
                            file_name=f"B2B_Leads_{hedef_kitle}_{country_code}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.warning("⚠️ Seçtiğiniz kriterlere uygun şirket bulunamadı. Anahtar kelimeleri değiştirmeyi deneyebilirsiniz.")
